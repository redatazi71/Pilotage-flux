"""Génère docs/cadrage_v4_kpis.xlsx avec les KPIs détaillés du cadrage v4.

Trois feuilles :
  - XL_4000_runs        : 5 scénarios × 4 doctrines (§24.2)
  - Random_1600_runs    : agrégat 4 doctrines (§24.6.2)
  - Decomposition_2x2   : matrice flux × event sourcing (§24.3, §24.6.3)

Usage : python docs/build_excel_kpis.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HERE = Path(__file__).resolve().parent
XLSX_PATH = HERE / "cadrage_v4_kpis.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="1F77B4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SCEN_FILL = PatternFill("solid", fgColor="E7EEF7")
TITLE_FONT = Font(bold=True, size=13, color="1F3A5F")


# --- Données XL (§24.2) ----------------------------------------------------

XL_KPIS = {
    "baseline_xl": {
        "lead_time": {"of": (8.65, 0.14), "flux": (4.84, 0.12),
                       "of_event": (8.34, 0.11), "event": (4.84, 0.12)},
        "wip":       {"of": (8.18, 0.14), "flux": (5.40, 0.11),
                       "of_event": (7.86, 0.11), "event": (5.40, 0.11)},
        "recalc":    {"of": 5.0, "flux": 5.0, "of_event": 2.0, "event": 2.0},
        "nervosity": {"of": 0.250, "flux": 0.250, "of_event": 0.100, "event": 0.100},
        "detections":{"of": 0, "flux": 0, "of_event": 127, "event": 96},
        "causes":    {"of": 0, "flux": 0, "of_event": 381, "event": 288},
        "cost":      {"of": (45067, 893), "flux": (32098, 140),
                       "of_event": (39274, 1498), "event": (32098, 140)},
    },
    "stress_double_breakdown_xl": {
        "lead_time": {"of": (8.81, 0.06), "flux": (5.00, 0.10),
                       "of_event": (8.27, 0.05), "event": (4.88, 0.00)},
        "wip":       {"of": (8.40, 0.10), "flux": (5.62, 0.10),
                       "of_event": (7.94, 0.08), "event": (5.40, 0.00)},
        "recalc":    {"of": 6.0, "flux": 6.0, "of_event": 2.5, "event": 2.5},
        "nervosity": {"of": 0.136, "flux": 0.136, "of_event": 0.045, "event": 0.045},
        "detections":{"of": 0, "flux": 0, "of_event": 178, "event": 142},
        "causes":    {"of": 0, "flux": 0, "of_event": 534, "event": 426},
        "cost":      {"of": (48586, 1688), "flux": (30251, 2093),
                       "of_event": (35890, 1721), "event": (27590, 0)},
    },
    "stress_cascade_nc_xl": {
        "lead_time": {"of": (8.22, 0.08), "flux": (4.62, 0.06),
                       "of_event": (8.18, 0.07), "event": (4.62, 0.06)},
        "wip":       {"of": (8.05, 0.10), "flux": (5.30, 0.08),
                       "of_event": (8.00, 0.10), "event": (5.30, 0.08)},
        "recalc":    {"of": 5.0, "flux": 5.0, "of_event": 2.5, "event": 2.5},
        "nervosity": {"of": 0.250, "flux": 0.250, "of_event": 0.050, "event": 0.050},
        "detections":{"of": 0, "flux": 0, "of_event": 56, "event": 42},
        "causes":    {"of": 0, "flux": 0, "of_event": 168, "event": 126},
        "cost":      {"of": (34125, 234), "flux": (27845, 112),
                       "of_event": (34056, 138), "event": (27718, 56)},
    },
    "stress_demand_spike_xl": {
        "lead_time": {"of": (10.19, 0.20), "flux": (6.36, 0.15),
                       "of_event": (10.19, 0.20), "event": (6.36, 0.15)},
        "wip":       {"of": (12.10, 0.30), "flux": (8.20, 0.20),
                       "of_event": (12.10, 0.30), "event": (8.20, 0.20)},
        "recalc":    {"of": 5.5, "flux": 5.5, "of_event": 2.0, "event": 2.0},
        "nervosity": {"of": 0.273, "flux": 0.273, "of_event": 0.091, "event": 0.091},
        "detections":{"of": 0, "flux": 0, "of_event": 24, "event": 18},
        "causes":    {"of": 0, "flux": 0, "of_event": 72, "event": 54},
        "cost":      {"of": (47173, 548), "flux": (41680, 1620),
                       "of_event": (47173, 548), "event": (41680, 1620)},
    },
    "stress_multi_contract_overload": {
        "lead_time": {"of": (4.50, 0.00), "flux": (2.00, 0.00),
                       "of_event": (4.50, 0.00), "event": (2.00, 0.00)},
        "wip":       {"of": (20.00, 0.00), "flux": (10.14, 0.00),
                       "of_event": (20.00, 0.00), "event": (10.14, 0.00)},
        "recalc":    {"of": 4.0, "flux": 4.0, "of_event": 1.5, "event": 1.5},
        "nervosity": {"of": 0.286, "flux": 0.286, "of_event": 0.143, "event": 0.143},
        "detections":{"of": 0, "flux": 0, "of_event": 36, "event": 27},
        "causes":    {"of": 0, "flux": 0, "of_event": 108, "event": 81},
        "cost":      {"of": (21391, 0), "flux": (9942, 0),
                       "of_event": (20104, 745), "event": (9942, 0)},
    },
}
DOCTRINES = ["of", "flux", "of_event", "event"]
DOCTRINE_LABELS = {"of": "OF", "flux": "FLUX",
                   "of_event": "OF+EVENT", "event": "EVENT"}


# --- Données Random (§24.6.2) ----------------------------------------------

RANDOM_KPIS = {
    "of":       {"lead_time": (8.61, 1.27), "wip": (14.77, 5.08),
                 "cost": (201774, 58896), "recalc": 7.0,
                 "nervosity": 0.350, "detections": 0.0, "causes": 0.0,
                 "delta_vs_of": 0},
    "flux":     {"lead_time": (4.80, 0.94), "wip": (9.20, 3.78),
                 "cost": (162481, 52847), "recalc": 7.0,
                 "nervosity": 0.350, "detections": 0.0, "causes": 0.0,
                 "delta_vs_of": -39293},
    "of_event": {"lead_time": (8.52, 1.25), "wip": (14.68, 5.08),
                 "cost": (196134, 57430), "recalc": 1.8,
                 "nervosity": 0.090, "detections": 213.3, "causes": 640.0,
                 "delta_vs_of": -5640},
    "event":    {"lead_time": (4.69, 0.94), "wip": (9.12, 3.81),
                 "cost": (157780, 53929), "recalc": 1.8,
                 "nervosity": 0.090, "detections": 134.7, "causes": 404.0,
                 "delta_vs_of": -43994},
}


# --- Helpers --------------------------------------------------------------

def _style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)


def _autofit(ws, min_width: int = 11, max_width: int = 32) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = min_width
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if length > width:
                    width = length
        ws.column_dimensions[letter].width = min(width + 2, max_width)


def _fmt_pair(value: tuple[float, float] | float, precision: int = 2) -> str:
    if isinstance(value, tuple):
        m, s = value
        return f"{m:.{precision}f} ± {s:.{precision}f}"
    return f"{value:.{precision}f}" if isinstance(value, float) else str(value)


def _fmt_money(value: tuple[float, float] | float) -> str:
    if isinstance(value, tuple):
        m, s = value
        return f"{int(m):,} ± {int(s):,} €".replace(",", " ")
    return f"{int(value):,} €".replace(",", " ")


# --- Sheets ---------------------------------------------------------------

def sheet_xl(wb: Workbook) -> None:
    ws = wb.create_sheet("XL_4000_runs")
    ws.cell(1, 1, "§24.2 — Étude XL 4 000 runs (5 scénarios × 4 doctrines × 200 seeds)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    headers = ["Scénario", "Doctrine", "Lead time (j)", "WIP", "Recalc APS",
               "Nervosité", "Détections", "Causes", "Coût total"]
    for j, h in enumerate(headers, start=1):
        _style_header(ws.cell(3, j, h))

    row = 4
    for scen, kpis in XL_KPIS.items():
        first_row_scen = row
        for d in DOCTRINES:
            ws.cell(row, 1, scen)
            ws.cell(row, 2, DOCTRINE_LABELS[d])
            ws.cell(row, 3, _fmt_pair(kpis["lead_time"][d]))
            ws.cell(row, 4, _fmt_pair(kpis["wip"][d]))
            ws.cell(row, 5, kpis["recalc"][d])
            ws.cell(row, 6, _fmt_pair(kpis["nervosity"][d], precision=3))
            ws.cell(row, 7, kpis["detections"][d])
            ws.cell(row, 8, kpis["causes"][d])
            ws.cell(row, 9, _fmt_money(kpis["cost"][d]))
            row += 1
        ws.merge_cells(start_row=first_row_scen, start_column=1,
                        end_row=row - 1, end_column=1)
        sc = ws.cell(first_row_scen, 1)
        sc.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        sc.fill = SCEN_FILL
        sc.font = Font(bold=True)

    _autofit(ws)


def sheet_random(wb: Workbook) -> None:
    ws = wb.create_sheet("Random_1600_runs")
    ws.cell(1, 1, "§24.6.2 — Étude Random 1 600 runs (20 fixtures × 20 scénarios × 4 doctrines)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    headers = ["Doctrine", "Lead time (j)", "WIP", "Recalc APS", "Nervosité",
               "Détections moy.", "Causes moy.", "Coût total", "Δ vs OF"]
    for j, h in enumerate(headers, start=1):
        _style_header(ws.cell(3, j, h))

    row = 4
    for d in DOCTRINES:
        k = RANDOM_KPIS[d]
        ws.cell(row, 1, DOCTRINE_LABELS[d])
        ws.cell(row, 2, _fmt_pair(k["lead_time"]))
        ws.cell(row, 3, _fmt_pair(k["wip"]))
        ws.cell(row, 4, k["recalc"])
        ws.cell(row, 5, _fmt_pair(k["nervosity"], precision=3))
        ws.cell(row, 6, k["detections"])
        ws.cell(row, 7, k["causes"])
        ws.cell(row, 8, _fmt_money(k["cost"]))
        delta = k["delta_vs_of"]
        ws.cell(row, 9, f"{delta:+,} €".replace(",", " ") if delta else "0 (réf)")
        if delta < 0:
            ws.cell(row, 9).font = Font(bold=True, color="2CA02C")
        row += 1

    _autofit(ws)


def sheet_decomposition(wb: Workbook) -> None:
    ws = wb.create_sheet("Decomposition_2x2")
    ws.cell(1, 1, "§24.3 + §24.6.3 — Décomposition 2×2 : apport flux × apport event sourcing").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    # --- Par scénario XL (§24.3)
    ws.cell(3, 1, "§24.3 — Décomposition par scénario XL (Δ coût € vs OF)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)

    headers = ["Scénario", "OF (réf €)", "FLUX seul Δ €",
               "OF+EVENT seul Δ €", "EVENT combiné Δ €"]
    for j, h in enumerate(headers, start=1):
        _style_header(ws.cell(4, j, h))

    decomp_xl = [
        ("baseline_xl", 45067, -12969, -5793, -12969),
        ("stress_double_breakdown_xl", 48586, -18336, -12697, -20996),
        ("stress_cascade_nc_xl", 34125, -6280, -69, -6407),
        ("stress_demand_spike_xl", 47173, -5492, 0, -5492),
        ("stress_multi_contract_overload", 21391, -11449, -1287, -11449),
    ]
    row = 5
    for name, of, flux, ofev, ev in decomp_xl:
        ws.cell(row, 1, name)
        ws.cell(row, 2, f"{of:,} €".replace(",", " "))
        ws.cell(row, 3, f"{flux:+,} €".replace(",", " "))
        ws.cell(row, 4, f"{ofev:+,} €".replace(",", " "))
        ws.cell(row, 5, f"{ev:+,} €".replace(",", " "))
        for col in (3, 4, 5):
            if ws.cell(row, col).value.startswith("-"):
                ws.cell(row, col).font = Font(color="2CA02C")
        row += 1

    # --- Matrice 2×2 Random (§24.6.3)
    row += 2
    ws.cell(row, 1, "§24.6.3 — Matrice 2×2 globale Random (1 600 runs)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    _style_header(ws.cell(row, 1, ""))
    _style_header(ws.cell(row, 2, "Flux ✗"))
    _style_header(ws.cell(row, 3, "Flux ✓"))
    row += 1

    ws.cell(row, 1, "Event ✗").font = Font(bold=True)
    ws.cell(row, 2, "0 (réf)")
    ws.cell(row, 3, "−39 293 €")
    ws.cell(row, 3).font = Font(bold=True, color="2CA02C")
    row += 1
    ws.cell(row, 1, "Event ✓").font = Font(bold=True)
    ws.cell(row, 2, "−5 640 €")
    ws.cell(row, 2).font = Font(color="2CA02C")
    ws.cell(row, 3, "−43 994 €")
    ws.cell(row, 3).font = Font(bold=True, color="2CA02C", size=12)

    # --- Additivité (§24.6.4)
    row += 2
    ws.cell(row, 1, "§24.6.4 — Additivité quasi-parfaite des deux apports").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    for label, value in [
        ("FLUX seul", -39293),
        ("OF+EVENT seul", -5640),
        ("Sommé naïvement", -44933),
        ("EVENT combiné mesuré", -43994),
        ("Sub-additivité (interaction)", 939),
    ]:
        ws.cell(row, 1, label).font = Font(bold=True)
        ws.cell(row, 2, f"{value:+,} €".replace(",", " "))
        row += 1

    _autofit(ws)


def sheet_resilience(wb: Workbook) -> None:
    ws = wb.create_sheet("Resilience_856_runs")
    ws.cell(1, 1, "§24.8 — Analyse de résilience (856 runs : 256 distrib + 300 gradient + 300 cascade)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # §24.8.1 — Distributions
    ws.cell(3, 1, "§24.8.1 — Distributions de coût (256 runs)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)
    headers = ["Doctrine", "N", "Moyenne", "σ", "P50", "P75", "P95", "P99", "Max"]
    for j, h in enumerate(headers, start=1):
        _style_header(ws.cell(4, j, h))

    dist_rows = [
        ("OF", 64, 152028, 55230, 136213, 174280, 261855, 316851, 359912),
        ("FLUX", 64, 126940, 49975, 119951, 147237, 222570, 289395, 321560),
        ("OF+EVENT", 64, 146672, 56155, 135383, 167709, 276428, 310432, 342564),
        ("EVENT", 64, 121866, 50233, 112829, 141320, 220793, 286827, 314621),
    ]
    for i, (name, n, m, s, p50, p75, p95, p99, mx) in enumerate(dist_rows, start=5):
        ws.cell(i, 1, name).font = Font(bold=True)
        ws.cell(i, 2, n)
        for col, v in enumerate([m, s, p50, p75, p95, p99, mx], start=3):
            ws.cell(i, col, f"{v:,} €".replace(",", " "))
        if name == "EVENT":
            ws.cell(i, 7).font = Font(bold=True, color="2CA02C")
            ws.cell(i, 8).font = Font(bold=True, color="2CA02C")

    # §24.8.2 — Gradient
    row = 11
    ws.cell(row, 1, "§24.8.2 — Gradient d'intensité (300 runs) — coût moyen €").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    grad_headers = ["Intensité", "OF", "FLUX", "OF+EVENT", "EVENT"]
    for j, h in enumerate(grad_headers, start=1):
        _style_header(ws.cell(row, j, h))
    row += 1
    grad_data = [
        (0.5, 114049, 101007, 113191, 100566),
        (1.0, 117336,  98364, 113029,  97655),
        (1.5, 119776,  99664, 113464,  99437),
        (2.0, 117605,  99990, 113980, 100594),
        (2.5, 114925, 100969, 113484,  99499),
    ]
    for intensity, *vals in grad_data:
        ws.cell(row, 1, f"x{intensity}").font = Font(bold=True)
        for col, v in enumerate(vals, start=2):
            ws.cell(row, col, f"{v:,} €".replace(",", " "))
        row += 1
    row += 1
    ws.cell(row, 1, "Note : courbe plate → ces variations d'intensité ne discriminent pas.")
    ws.cell(row, 1).font = Font(italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

    # §24.8.3 — Cascade
    row += 2
    ws.cell(row, 1, "§24.8.3 — Cascade de pannes simultanées (300 runs) — coût moyen €").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    casc_headers = ["Pannes simultanées", "OF", "FLUX", "OF+EVENT", "EVENT"]
    for j, h in enumerate(casc_headers, start=1):
        _style_header(ws.cell(row, j, h))
    row += 1
    casc_data = [
        (1, 107116, 68524, 101392, 67198),
        (2, 112506, 72831, 103452, 68398),
        (3, 120826, 75345, 105518, 69145),
        (4, 127923, 78276, 107685, 70087),
        (5, 131954, 80370, 110331, 70247),
    ]
    for n_bd, *vals in casc_data:
        ws.cell(row, 1, f"{n_bd}").font = Font(bold=True)
        for col, v in enumerate(vals, start=2):
            ws.cell(row, col, f"{v:,} €".replace(",", " "))
        ws.cell(row, 5).font = Font(bold=True, color="2CA02C")
        row += 1

    # MTTR
    row += 1
    ws.cell(row, 1, "Time-to-recover (jours)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    for j, h in enumerate(casc_headers, start=1):
        _style_header(ws.cell(row, j, h))
    row += 1
    mttr_data = [
        (1, 5.8, 3.0, 5.7, 2.9),
        (2, 5.9, 3.5, 5.7, 3.0),
        (3, 5.9, 3.9, 5.7, 3.1),
        (4, 5.7, 4.9, 5.7, 3.5),
        (5, 5.5, 5.1, 5.7, 3.5),
    ]
    for n_bd, *vals in mttr_data:
        ws.cell(row, 1, f"{n_bd}").font = Font(bold=True)
        for col, v in enumerate(vals, start=2):
            ws.cell(row, col, f"{v:.1f} j")
        ws.cell(row, 5).font = Font(bold=True, color="2CA02C")
        row += 1

    # Sensibilité
    row += 1
    ws.cell(row, 1, "Sensibilité au choc — Δ relatif 1→5 pannes").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1
    sens_data = [("OF", "+23.2 %"), ("FLUX", "+17.3 %"),
                 ("OF+EVENT", "+8.8 %"), ("EVENT", "+4.5 %")]
    for name, delta in sens_data:
        ws.cell(row, 1, name).font = Font(bold=True)
        ws.cell(row, 2, delta)
        if name == "EVENT":
            ws.cell(row, 2).font = Font(bold=True, color="2CA02C", size=12)
        row += 1

    _autofit(ws)


def sheet_extension(wb: Workbook) -> None:
    """§24.8.5 + §24.10 — Point de rupture + Matrice paires."""
    ws = wb.create_sheet("Paires_400_runs")
    ws.cell(1, 1, "§24.10 — Matrice 5×5 paires de domaines (400 runs)").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    domains = ["Appro", "Logi", "Qual", "Prod", "Dem"]
    doctrines = [("OF", "of"), ("FLUX", "flux"),
                 ("OF+EVENT", "of_event"), ("EVENT", "event")]

    # Amplification matrices — données brutes (depuis cadrage_v4_resilience_ext_data.md)
    amplification = {
        "of":       [[1.00, 1.49, 1.00, 1.05, 1.03],
                     [1.49, 1.20, 1.49, 0.74, 1.49],
                     [1.00, 1.49, 1.00, 1.05, 1.02],
                     [1.05, 0.74, 1.05, 1.05, 1.07],
                     [1.03, 1.49, 1.02, 1.07, 1.05]],
        "flux":     [[1.00, 1.36, 1.00, 1.03, 1.20],
                     [1.36, 1.70, 1.36, 1.08, 2.20],
                     [1.00, 1.36, 1.00, 1.01, 1.19],
                     [1.03, 1.08, 1.01, 1.03, 1.23],
                     [1.20, 2.20, 1.19, 1.23, 1.05]],
        "of_event": [[1.00, 1.13, 1.00, 1.01, 1.03],
                     [1.13, 1.23, 1.13, 0.87, 1.15],
                     [1.00, 1.13, 1.00, 1.01, 1.02],
                     [1.01, 0.87, 1.01, 1.01, 1.04],
                     [1.03, 1.15, 1.02, 1.04, 1.05]],
        "event":    [[1.00, 0.99, 1.00, 1.01, 1.20],
                     [0.99, 1.52, 0.99, 1.01, 1.16],
                     [1.00, 0.99, 1.00, 1.01, 1.19],
                     [1.01, 1.01, 1.01, 1.01, 1.21],
                     [1.20, 1.16, 1.19, 1.21, 1.05]],
    }
    recovery = {
        "of":       [[5.8, 5.6, 5.8, 6.6, 5.4],
                     [5.6, 4.8, 5.6, 5.8, 5.8],
                     [5.8, 5.6, 5.8, 6.4, 5.6],
                     [6.6, 5.8, 6.4, 6.6, 5.6],
                     [5.4, 5.8, 5.6, 5.6, 5.2]],
        "flux":     [[2.4, 2.8, 2.4, 4.0, 6.8],
                     [2.8, 3.0, 2.8, 5.0, 5.0],
                     [2.4, 2.8, 2.4, 4.0, 5.0],
                     [4.0, 5.0, 4.0, 5.0, 5.8],
                     [6.8, 5.0, 5.0, 5.8, 6.4]],
        "of_event": [[5.8, 6.0, 5.8, 6.0, 5.4],
                     [6.0, 5.8, 6.0, 6.0, 6.2],
                     [5.8, 6.0, 5.8, 5.8, 5.6],
                     [6.0, 6.0, 5.8, 6.0, 5.8],
                     [5.4, 6.2, 5.6, 5.8, 5.2]],
        "event":    [[2.4, 2.8, 2.4, 2.4, 6.8],
                     [2.8, 2.8, 2.8, 3.0, 5.2],
                     [2.4, 2.8, 2.4, 2.4, 5.0],
                     [2.4, 3.0, 2.4, 2.4, 5.6],
                     [6.8, 5.2, 5.0, 5.6, 6.4]],
    }

    row = 3
    ws.cell(row, 1, "§24.10.A — Amplification de coût (>1 = sur-coût)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    for label, key in doctrines:
        ws.cell(row, 1, label).font = Font(bold=True, color="1F3A5F")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        _style_header(ws.cell(row, 1, ""))
        for j, d in enumerate(domains, start=2):
            _style_header(ws.cell(row, j, d))
        row += 1
        m = amplification[key]
        for i, d in enumerate(domains):
            ws.cell(row, 1, d).font = Font(bold=True)
            for j, val in enumerate(m[i], start=2):
                cell = ws.cell(row, j, val)
                cell.number_format = "0.00"
                if val >= 1.50:
                    cell.font = Font(bold=True, color="C00000")
                elif val >= 1.20:
                    cell.font = Font(color="C04500")
                elif val < 0.95:
                    cell.font = Font(color="2CA02C")
            row += 1
        row += 1

    row += 1
    ws.cell(row, 1, "§24.10.B — Time-to-recover par paire (jours)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    for label, key in doctrines:
        ws.cell(row, 1, label).font = Font(bold=True, color="1F3A5F")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        _style_header(ws.cell(row, 1, ""))
        for j, d in enumerate(domains, start=2):
            _style_header(ws.cell(row, j, d))
        row += 1
        m = recovery[key]
        for i, d in enumerate(domains):
            ws.cell(row, 1, d).font = Font(bold=True)
            for j, val in enumerate(m[i], start=2):
                cell = ws.cell(row, j, val)
                cell.number_format = "0.0"
                if val >= 6.0:
                    cell.font = Font(bold=True, color="C00000")
                elif val <= 3.0:
                    cell.font = Font(bold=True, color="2CA02C")
            row += 1
        row += 1

    # Point de rupture
    row += 1
    ws.cell(row, 1, "§24.8.5 — Point de rupture (cascade poussée)").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    bp_headers = ["N pannes", "OF €", "FLUX €", "OF+EVENT €", "EVENT €",
                  "Dispo OF", "Dispo EVENT"]
    for j, h in enumerate(bp_headers, start=1):
        _style_header(ws.cell(row, j, h))
    row += 1
    bp_data = [
        (6, 142517, 85556, 119142, 74483, "99.5%", "94.4%"),
        (8, 151406, 89425, 122605, 76327, "99.5%", "94.4%"),
        (10, 151406, 89425, 122605, 76327, "99.5%", "94.4%"),
        (12, 151406, 89425, 122605, 76327, "99.5%", "94.4%"),
        (15, 151406, 89425, 122605, 76327, "99.5%", "94.4%"),
    ]
    for n_bd, of_c, flux_c, ofe_c, ev_c, of_a, ev_a in bp_data:
        ws.cell(row, 1, n_bd).font = Font(bold=True)
        for col, v in enumerate([of_c, flux_c, ofe_c, ev_c], start=2):
            ws.cell(row, col, f"{v:,} €".replace(",", " "))
        ws.cell(row, 6, of_a)
        ws.cell(row, 7, ev_a)
        ws.cell(row, 5).font = Font(bold=True, color="2CA02C")
        row += 1

    _autofit(ws)


def build_xlsx() -> None:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    sheet_xl(wb)
    sheet_random(wb)
    sheet_decomposition(wb)
    sheet_resilience(wb)
    sheet_extension(wb)
    wb.save(XLSX_PATH)
    print(f"Excel KPIs généré : {XLSX_PATH}")


if __name__ == "__main__":
    build_xlsx()
